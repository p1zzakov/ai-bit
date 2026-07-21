from __future__ import annotations

import json
import re
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, PlainTextResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

VERSION = "5.1.0"
ROOT = Path("/app/artifacts/onec-requirements")
PROJECTS = ROOT / "projects"
UPLOADS = ROOT / "uploads"
for folder in (PROJECTS, UPLOADS):
    folder.mkdir(parents=True, exist_ok=True)

router = APIRouter(prefix="/onec-requirements", tags=["1C Requirement Architect"])


class AnswerRequest(BaseModel):
    answers: dict[str, Any] = Field(default_factory=dict)


class PrototypeApproval(BaseModel):
    approved: bool
    comment: str = ""


def now_iso() -> str:
    return datetime.now(UTC).isoformat()


def project_path(project_id: str) -> Path:
    if not re.fullmatch(r"[a-f0-9]{32}", project_id):
        raise HTTPException(status_code=400, detail="Invalid project id")
    return PROJECTS / f"{project_id}.json"


def save(project: dict[str, Any]) -> dict[str, Any]:
    project["updated_at"] = now_iso()
    path = project_path(str(project["id"]))
    path.write_text(json.dumps(project, ensure_ascii=False, indent=2), encoding="utf-8")
    return project


def load(project_id: str) -> dict[str, Any]:
    path = project_path(project_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    return json.loads(path.read_text(encoding="utf-8"))


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_excel(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets[:10]:
        max_row = min(worksheet.max_row or 0, 200)
        max_col = min(worksheet.max_column or 0, 80)
        rows: list[list[Any]] = []
        formulas: list[dict[str, str]] = []
        nonempty_rows: list[int] = []
        for r in range(1, max_row + 1):
            row: list[Any] = []
            has_data = False
            for c in range(1, max_col + 1):
                cell = worksheet.cell(r, c)
                value = clean(cell.value)
                if value not in (None, ""):
                    has_data = True
                if isinstance(value, str) and value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": value})
                row.append(value)
            if has_data:
                nonempty_rows.append(r)
                rows.append(row)
        header_row = rows[0] if rows else []
        headers = [str(x).strip() if x not in (None, "") else f"Колонка {i + 1}" for i, x in enumerate(header_row)]
        samples = rows[1:11] if len(rows) > 1 else []
        sheets.append({
            "name": worksheet.title,
            "dimensions": {"rows": worksheet.max_row, "columns": worksheet.max_column},
            "headers": headers,
            "sample_rows": samples,
            "formulas": formulas[:100],
            "merged_ranges": [str(x) for x in list(worksheet.merged_cells.ranges)[:100]],
            "nonempty_rows": nonempty_rows[:200],
        })
    return {"format": "xlsx", "sheets": sheets, "sheet_count": len(sheets)}


def parse_text(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")[:100_000]
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return {"format": "text", "lines": lines[:500], "preview": text[:5000]}


def parse_upload(path: Path) -> dict[str, Any]:
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return parse_excel(path)
    if suffix in {".txt", ".csv", ".md"}:
        return parse_text(path)
    return {"format": suffix.lstrip(".") or "unknown", "status": "stored_not_parsed", "note": "Файл сохранён. Автоматический разбор этого формата будет добавлен позднее."}


def inferred_headers(parsed: dict[str, Any]) -> list[str]:
    sheets = parsed.get("sheets") if isinstance(parsed.get("sheets"), list) else []
    if not sheets:
        return []
    return [str(x) for x in sheets[0].get("headers", []) if str(x).strip()]


def build_questions(project: dict[str, Any]) -> list[dict[str, Any]]:
    headers = inferred_headers(project.get("source_analysis", {}))
    questions: list[dict[str, Any]] = [
        {"id": "business_goal", "section": "Назначение", "question": "Для чего нужен этот отчёт и какое решение принимается по его результату?", "required": True, "type": "textarea"},
        {"id": "users", "section": "Пользователи", "question": "Кто будет формировать и использовать отчёт?", "required": True, "type": "text"},
        {"id": "source_process", "section": "Источник", "question": "По какому процессу нужны данные: продажи, заказы, отгрузки, оплаты, закупки, склад или другое?", "required": True, "type": "text"},
        {"id": "period_logic", "section": "Период", "question": "За какой период строится отчёт и какая дата считается основной?", "required": True, "type": "textarea"},
        {"id": "filters", "section": "Отборы", "question": "Какие обязательные фильтры нужны пользователю?", "required": False, "type": "textarea"},
        {"id": "exclusions", "section": "Исключения", "question": "Какие документы, статусы или строки нельзя включать?", "required": False, "type": "textarea"},
        {"id": "totals", "section": "Итоги", "question": "Какие итоги, группировки и расчётные показатели должны быть в отчёте?", "required": True, "type": "textarea"},
        {"id": "accuracy", "section": "Приёмка", "question": "На каком известном примере можно проверить, что отчёт посчитан правильно?", "required": True, "type": "textarea"},
        {"id": "refresh", "section": "Эксплуатация", "question": "Как часто отчёт формируется и какой объём данных ожидается?", "required": False, "type": "text"},
        {"id": "delivery", "section": "Форма", "question": "Нужен внешний отчёт, встроенный отчёт, печатная форма или выгрузка в Excel?", "required": True, "type": "text"},
    ]
    if headers:
        questions.insert(4, {"id": "columns_confirm", "section": "Макет", "question": f"Подтвердите назначение колонок: {', '.join(headers[:25])}. Какие из них обязательны, а какие примерные?", "required": True, "type": "textarea"})
    joined = " ".join(headers).casefold()
    if any(token in joined for token in ("прибыл", "марж", "себестоим")):
        questions.append({"id": "profit_method", "section": "Расчёт", "question": "Прибыль должна быть плановой или фактической? На каком этапе появляется себестоимость?", "required": True, "type": "textarea"})
    if any(token in joined for token in ("остат", "склад")):
        questions.append({"id": "stock_moment", "section": "Расчёт", "question": "Остаток нужен текущий или на выбранную дату? Учитывать резерв и ожидаемое поступление?", "required": True, "type": "textarea"})
    return questions


def find_latest_mcp_snapshot() -> dict[str, Any]:
    candidates = [
        Path("/app/artifacts/external-sources/latest.json"),
        Path("/app/artifacts/bitrix-onec-integration/latest.json"),
    ]
    candidates.extend(sorted(Path("/app/artifacts").glob("**/latest*.json"), key=lambda p: p.stat().st_mtime, reverse=True)[:20])
    for path in candidates:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        text = json.dumps(payload, ensure_ascii=False).casefold()
        if "mcp_1c" in text or "onec_profile" in text or "управлениепредприятием" in text:
            return {"source": str(path), "payload": payload}
    return {"source": None, "payload": {}}


def recursive_strings(node: Any) -> list[str]:
    result: list[str] = []
    if isinstance(node, dict):
        for key, value in node.items():
            result.append(str(key))
            result.extend(recursive_strings(value))
    elif isinstance(node, list):
        for value in node:
            result.extend(recursive_strings(value))
    elif isinstance(node, (str, int, float)):
        result.append(str(node))
    return result


def assess(project: dict[str, Any]) -> dict[str, Any]:
    answers = project.get("answers", {})
    headers = inferred_headers(project.get("source_analysis", {}))
    snapshot = find_latest_mcp_snapshot()
    corpus = "\n".join(recursive_strings(snapshot.get("payload", {})))
    corpus_lower = corpus.casefold()
    requirements_text = " ".join(str(x) for x in answers.values()).casefold() + " " + " ".join(headers).casefold()

    warnings: list[dict[str, Any]] = []
    if any(x in requirements_text for x in ("прибыл", "марж", "себестоим")):
        warnings.append({"severity": "high", "code": "CALC-PROFIT-01", "title": "Необходимо определить момент возникновения себестоимости", "explanation": "Фактическая прибыль обычно не может быть корректно рассчитана только по заказу. Требуется связать реализацию и движения регистров себестоимости.", "resolution": "Зафиксировать: плановая прибыль по заказу либо фактическая прибыль после реализации."})
    if "остат" in requirements_text and not answers.get("stock_moment"):
        warnings.append({"severity": "high", "code": "CALC-STOCK-01", "title": "Не определён момент расчёта остатка", "explanation": "Текущий остаток и остаток на дату — разные алгоритмы и источники.", "resolution": "Указать дату среза, склад, учёт резерва и доступного остатка."})
    if not answers.get("accuracy"):
        warnings.append({"severity": "high", "code": "ACCEPT-01", "title": "Нет контрольного примера для приёмки", "explanation": "Без эталонного примера невозможно доказать корректность расчёта.", "resolution": "Подготовить минимум один период и набор документов с заранее проверенным результатом."})
    if not snapshot.get("source"):
        warnings.append({"severity": "medium", "code": "MCP-01", "title": "Не найден актуальный снимок метаданных 1С", "explanation": "Источники данных не подтверждены по реальной конфигурации.", "resolution": "Запустить сбор MCP 1С и повторить анализ проекта ТЗ."})

    object_candidates = []
    keyword_map = {
        "заказ": ("ЗаказКлиента", "Документ"),
        "продаж": ("РеализацияТоваровУслуг", "Документ"),
        "отгруз": ("РеализацияТоваровУслуг", "Документ"),
        "клиент": ("Партнеры / Контрагенты", "Справочник"),
        "товар": ("Номенклатура", "Справочник"),
        "склад": ("Склады", "Справочник"),
        "остат": ("ТоварыНаСкладах", "Регистр накопления"),
        "оплат": ("БезналичныеПлатежи / Оплата", "Документ/регистр"),
        "себестоим": ("СебестоимостьТоваров", "Регистр накопления"),
    }
    for keyword, (name, kind) in keyword_map.items():
        if keyword in requirements_text:
            object_candidates.append({"name": name, "type": kind, "confirmed_in_snapshot": name.casefold().split(" /")[0] in corpus_lower, "reason": f"Требование содержит смысловой маркер «{keyword}»"})

    if warnings and any(x["severity"] == "high" for x in warnings):
        feasibility = "requires_clarification"
    elif snapshot.get("source"):
        feasibility = "feasible_with_external_report"
    else:
        feasibility = "insufficient_evidence"

    columns = [{"order": i + 1, "name": header, "purpose": "Из загруженного макета", "source": "Требует подтверждения", "calculation": "Без расчёта"} for i, header in enumerate(headers)]
    return {
        "status": feasibility,
        "mcp_evidence": {"available": bool(snapshot.get("source")), "source": snapshot.get("source")},
        "candidate_objects": object_candidates,
        "warnings": warnings,
        "prototype": {
            "title": project.get("title"),
            "columns": columns,
            "filters": str(answers.get("filters") or "Не определены"),
            "grouping_and_totals": str(answers.get("totals") or "Не определены"),
            "sample_rows": (project.get("source_analysis", {}).get("sheets") or [{}])[0].get("sample_rows", [])[:5] if project.get("source_analysis", {}).get("sheets") else [],
            "status": "awaiting_approval",
        },
    }


def build_spec(project: dict[str, Any]) -> dict[str, Any]:
    analysis = project.get("analysis") or assess(project)
    answers = project.get("answers", {})
    prototype = analysis.get("prototype", {})
    columns = prototype.get("columns", [])
    warnings = analysis.get("warnings", [])
    candidates = analysis.get("candidate_objects", [])
    return {
        "document": "Техническое задание на разработку отчёта в 1С",
        "version": project.get("revision", 1),
        "project_id": project.get("id"),
        "status": "approved_for_development" if project.get("prototype_approval", {}).get("approved") else "draft",
        "configuration_evidence": analysis.get("mcp_evidence"),
        "business": {
            "goal": answers.get("business_goal", project.get("description", "")),
            "users": answers.get("users", "Не определены"),
            "result": answers.get("totals", "Не определён"),
            "delivery": answers.get("delivery", "Внешний отчёт"),
        },
        "functional_requirements": {
            "period": answers.get("period_logic", "Не определён"),
            "filters": answers.get("filters", "Не определены"),
            "exclusions": answers.get("exclusions", "Не определены"),
            "columns": columns,
            "grouping_and_totals": answers.get("totals", "Не определены"),
        },
        "technical_design": {
            "recommended_implementation": "Внешний отчёт на СКД; изменение типовой конфигурации не выполнять без отдельного согласования.",
            "candidate_data_sources": candidates,
            "performance": "Обязательные отборы по периоду; исключить полное чтение регистров без ограничений; проверить план запроса на рабочем объёме.",
            "security": "Использовать права текущего пользователя; не расширять доступ к данным отчётом.",
            "read_only": True,
        },
        "risks_and_open_questions": warnings,
        "acceptance": {
            "reference_example": answers.get("accuracy", "Не предоставлен"),
            "tests": [
                "Проверить формирование на согласованном контрольном периоде.",
                "Сверить каждую итоговую сумму с первичными документами и стандартными отчётами 1С.",
                "Проверить пустые значения, отменённые и непроведённые документы.",
                "Проверить права минимум двух пользовательских ролей.",
                "Измерить время формирования на рабочем объёме данных.",
                "Проверить выгрузку в Excel без потери типов, итогов и группировок.",
            ],
        },
        "constraints": ["Не изменять типовую конфигурацию без отдельного решения.", "Не выполнять запись данных из отчёта.", "Не считать неподтверждённые источники фактом."],
    }


def spec_markdown(spec: dict[str, Any]) -> str:
    business = spec["business"]
    functional = spec["functional_requirements"]
    technical = spec["technical_design"]
    lines = [
        f"# {spec['document']}", "", f"**Версия:** {spec['version']}", f"**Статус:** {spec['status']}", "",
        "## 1. Назначение", str(business.get("goal") or "Не определено"), "",
        "## 2. Пользователи", str(business.get("users") or "Не определены"), "",
        "## 3. Функциональные требования", f"- Период: {functional.get('period')}", f"- Отборы: {functional.get('filters')}", f"- Исключения: {functional.get('exclusions')}", f"- Группировки и итоги: {functional.get('grouping_and_totals')}", "",
        "### Колонки отчёта",
    ]
    for column in functional.get("columns", []):
        lines.append(f"- {column.get('order')}. **{column.get('name')}** — источник: {column.get('source')}; расчёт: {column.get('calculation')}")
    lines.extend(["", "## 4. Предлагаемая реализация", technical.get("recommended_implementation", ""), "", "### Источники данных"])
    for source in technical.get("candidate_data_sources", []):
        lines.append(f"- {source.get('type')}: **{source.get('name')}**; подтверждено MCP: {source.get('confirmed_in_snapshot')}")
    lines.extend(["", "## 5. Риски и открытые вопросы"])
    for risk in spec.get("risks_and_open_questions", []):
        lines.append(f"- **{risk.get('severity', '').upper()} {risk.get('code')} — {risk.get('title')}**: {risk.get('explanation')} Исправление: {risk.get('resolution')}")
    lines.extend(["", "## 6. Критерии приёмки", f"Контрольный пример: {spec['acceptance'].get('reference_example')}"])
    for test in spec["acceptance"].get("tests", []):
        lines.append(f"- [ ] {test}")
    lines.extend(["", "## 7. Ограничения"])
    for constraint in spec.get("constraints", []):
        lines.append(f"- {constraint}")
    return "\n".join(lines) + "\n"


@router.get("", response_class=HTMLResponse)
def page() -> str:
    return REQUIREMENT_ARCHITECT_HTML


@router.get("/api/projects")
def list_projects() -> dict[str, Any]:
    rows = []
    for path in sorted(PROJECTS.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            project = json.loads(path.read_text(encoding="utf-8"))
            rows.append({key: project.get(key) for key in ("id", "title", "status", "created_at", "updated_at")})
        except Exception:
            continue
    return {"version": VERSION, "projects": rows}


@router.post("/api/projects")
async def create_project(title: str = Form(...), description: str = Form(""), file: UploadFile | None = File(None)) -> dict[str, Any]:
    project_id = uuid.uuid4().hex
    source_analysis: dict[str, Any] = {"format": "none", "note": "Файл не загружен"}
    upload_meta = None
    if file and file.filename:
        safe_name = re.sub(r"[^A-Za-zА-Яа-яЁё0-9._-]", "_", Path(file.filename).name)
        target = UPLOADS / f"{project_id}-{safe_name}"
        content = await file.read()
        if len(content) > 25 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File exceeds 25 MB")
        target.write_bytes(content)
        source_analysis = parse_upload(target)
        upload_meta = {"filename": file.filename, "stored_as": str(target), "size": len(content), "content_type": file.content_type}
    project = {
        "id": project_id,
        "version": VERSION,
        "revision": 1,
        "title": title.strip(),
        "description": description.strip(),
        "status": "interview_required",
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "upload": upload_meta,
        "source_analysis": source_analysis,
        "answers": {},
        "questions": [],
        "history": [{"at": now_iso(), "event": "project_created"}],
    }
    project["questions"] = build_questions(project)
    return save(project)


@router.get("/api/projects/{project_id}")
def get_project(project_id: str) -> dict[str, Any]:
    return load(project_id)


@router.post("/api/projects/{project_id}/answers")
def submit_answers(project_id: str, request: AnswerRequest) -> dict[str, Any]:
    project = load(project_id)
    project["answers"] = {**project.get("answers", {}), **request.answers}
    required = [q["id"] for q in project.get("questions", []) if q.get("required")]
    missing = [key for key in required if not str(project["answers"].get(key, "")).strip()]
    project["status"] = "interview_required" if missing else "ready_for_analysis"
    project["missing_required_answers"] = missing
    project["history"].append({"at": now_iso(), "event": "answers_updated", "missing": missing})
    return save(project)


@router.post("/api/projects/{project_id}/analyze")
def analyze_project(project_id: str) -> dict[str, Any]:
    project = load(project_id)
    project["analysis"] = assess(project)
    project["status"] = "prototype_review"
    project["history"].append({"at": now_iso(), "event": "analysis_completed", "result": project["analysis"]["status"]})
    return save(project)


@router.post("/api/projects/{project_id}/prototype-approval")
def approve_prototype(project_id: str, request: PrototypeApproval) -> dict[str, Any]:
    project = load(project_id)
    project["prototype_approval"] = {"approved": request.approved, "comment": request.comment, "at": now_iso()}
    project["status"] = "specification_ready" if request.approved else "prototype_revision_required"
    project["revision"] = int(project.get("revision", 1)) + (0 if request.approved else 1)
    project["specification"] = build_spec(project)
    project["history"].append({"at": now_iso(), "event": "prototype_approval", "approved": request.approved, "comment": request.comment})
    return save(project)


@router.get("/api/projects/{project_id}/spec")
def get_spec(project_id: str) -> dict[str, Any]:
    project = load(project_id)
    spec = project.get("specification") or build_spec(project)
    return spec


@router.get("/api/projects/{project_id}/spec.md", response_class=PlainTextResponse)
def get_spec_markdown(project_id: str) -> str:
    project = load(project_id)
    spec = project.get("specification") or build_spec(project)
    return spec_markdown(spec)


REQUIREMENT_ARCHITECT_HTML = r'''<!doctype html><html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Конструктор ТЗ для 1С — AI-BIT</title><style>
:root{color-scheme:light;--bg:#f5f6f8;--surface:#fff;--surface2:#fafbfc;--line:#e2e5e9;--text:#17191c;--muted:#70757d;--accent:#5e6ad2;--ok:#16835d;--warn:#a9670b;--bad:#b42318}*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--text);font:14px/1.5 Inter,Segoe UI,sans-serif}main{max-width:1500px;margin:auto;padding:28px}.head{display:flex;justify-content:space-between;gap:20px;align-items:flex-start;margin-bottom:18px}.eyebrow{font-size:11px;font-weight:800;letter-spacing:.12em;text-transform:uppercase;color:var(--accent)}h1{margin:4px 0;font-size:28px}.muted{color:var(--muted)}.grid{display:grid;grid-template-columns:340px minmax(0,1fr);gap:14px}.panel{background:#fff;border:1px solid var(--line);border-radius:12px;overflow:hidden;margin-bottom:12px}.ph{padding:13px 15px;border-bottom:1px solid var(--line);background:var(--surface2);display:flex;justify-content:space-between;align-items:center}.ph h2{margin:0;font-size:15px}.body{padding:15px}.project{padding:11px;border:1px solid var(--line);border-radius:9px;margin-bottom:8px;cursor:pointer}.project:hover,.project.active{border-color:var(--accent);background:#f7f7ff}.badge{display:inline-flex;border-radius:999px;padding:3px 8px;font-size:10px;font-weight:800;text-transform:uppercase;background:#eef0f2}.interview_required,.requires_clarification,.prototype_revision_required{background:#fff3dc;color:var(--warn)}.specification_ready,.feasible_with_external_report{background:#e7f6ef;color:var(--ok)}.insufficient_evidence{background:#fee4e2;color:var(--bad)}label{display:block;font-weight:700;margin:10px 0 5px}input,textarea{width:100%;border:1px solid var(--line);border-radius:8px;padding:9px 10px;font:inherit;background:#fff}textarea{min-height:80px;resize:vertical}.btn{border:1px solid var(--line);background:#fff;border-radius:8px;padding:9px 12px;cursor:pointer}.btn.primary{background:#202124;border-color:#202124;color:#fff}.btn.accent{background:var(--accent);border-color:var(--accent);color:#fff}.actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px}.step{padding:12px 14px;border-left:3px solid var(--line);margin-bottom:8px;background:#fafbfc}.step.done{border-color:var(--ok)}.question{padding:12px;border:1px solid var(--line);border-radius:9px;margin-bottom:9px}.question small{display:block;color:var(--muted);margin-bottom:5px}.warning{border:1px solid #f2d5a5;background:#fffaf0;border-radius:9px;padding:12px;margin-bottom:8px}.warning.high{border-color:#efb3ae;background:#fff6f5}.tablewrap{overflow:auto}table{width:100%;border-collapse:collapse}th,td{padding:9px 10px;border-bottom:1px solid #edf0f2;text-align:left;vertical-align:top}th{font-size:10px;text-transform:uppercase;color:var(--muted);background:#fbfcfd}.mono{font-family:ui-monospace,SFMono-Regular,Consolas,monospace;font-size:12px}.empty{padding:25px;text-align:center;color:var(--muted)}@media(max-width:900px){.grid{grid-template-columns:1fr}.head{display:block}}
</style></head><body><main><div class="head"><div><div class="eyebrow">AI-BIT Enterprise 5.1</div><h1>Конструктор ТЗ для 1С</h1><div class="muted">Из сырого Excel и разговора с пользователем — в проверенное, согласованное техническое задание.</div></div><button class="btn primary" id="newBtn">Новое ТЗ</button></div><div class="grid"><aside><div class="panel"><div class="ph"><h2>Проекты</h2><span id="count"></span></div><div class="body" id="projects"></div></div></aside><section id="workspace"><div class="panel"><div class="body empty">Создайте новый проект или выберите существующий.</div></div></section></div></main><script>
const q=s=>document.querySelector(s),esc=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));let current=null;const badge=v=>`<span class="badge ${esc(v)}">${esc(v)}</span>`;async function api(url,opt={}){const r=await fetch(url,{cache:'no-store',...opt});if(!r.ok)throw new Error(await r.text());return r.json()}async function list(){const d=await api('/onec-requirements/api/projects');q('#count').textContent=d.projects.length;q('#projects').innerHTML=d.projects.map(x=>`<div class="project ${current===x.id?'active':''}" data-id="${x.id}"><b>${esc(x.title)}</b><div>${badge(x.status)}</div><small class="muted">${esc(x.updated_at||'')}</small></div>`).join('')||'<div class="empty">Проектов пока нет.</div>';document.querySelectorAll('.project').forEach(x=>x.onclick=()=>openProject(x.dataset.id))}function newForm(){current=null;q('#workspace').innerHTML=`<div class="panel"><div class="ph"><h2>Новое техническое задание</h2></div><div class="body"><form id="create"><label>Название задачи</label><input name="title" required placeholder="Например: Отчёт по марже заказов"><label>Опишите пожелание своими словами</label><textarea name="description" placeholder="Что хотите получить и зачем"></textarea><label>Пример отчёта</label><input name="file" type="file" accept=".xlsx,.xlsm,.csv,.txt,.md"><div class="muted">Лучше всего загрузить Excel с желаемыми колонками и примерами строк.</div><div class="actions"><button class="btn accent">Создать и начать опрос</button></div></form></div></div>`;q('#create').onsubmit=async e=>{e.preventDefault();const p=await api('/onec-requirements/api/projects',{method:'POST',body:new FormData(e.target)});await list();openProject(p.id)}}function steps(p){const items=[['Исходный пример',!!p.source_analysis],['Интервью',p.status!=='interview_required'],['Проверка 1С',!!p.analysis],['Прототип',!!p.analysis?.prototype],['Согласование',!!p.prototype_approval],['Техническое ТЗ',!!p.specification]];return items.map((x,i)=>`<div class="step ${x[1]?'done':''}"><b>${i+1}. ${x[0]}</b><div class="muted">${x[1]?'Выполнено':'Ожидает'}</div></div>`).join('')}async function openProject(id){current=id;await list();const p=await api('/onec-requirements/api/projects/'+id);const questions=(p.questions||[]).map(x=>`<div class="question"><small>${esc(x.section)} ${x.required?'· обязательно':''}</small><b>${esc(x.question)}</b>${x.type==='textarea'?`<textarea data-answer="${x.id}">${esc(p.answers?.[x.id]||'')}</textarea>`:`<input data-answer="${x.id}" value="${esc(p.answers?.[x.id]||'')}">`}</div>`).join('');const warnings=(p.analysis?.warnings||[]).map(x=>`<div class="warning ${x.severity}"><b>${esc(x.code)} · ${esc(x.title)}</b><div>${esc(x.explanation)}</div><div><strong>Что сделать:</strong> ${esc(x.resolution)}</div></div>`).join('');const proto=p.analysis?.prototype;const spec=p.specification;q('#workspace').innerHTML=`<div class="panel"><div class="ph"><div><h2>${esc(p.title)}</h2>${badge(p.status)}</div><span class="mono">rev ${p.revision}</span></div><div class="body"><div class="grid" style="grid-template-columns:260px 1fr"><div>${steps(p)}</div><div><h3>Описание</h3><p>${esc(p.description||'—')}</p><h3>Интервью с заказчиком</h3>${questions}<div class="actions"><button class="btn accent" id="saveAnswers">Сохранить ответы</button><button class="btn primary" id="analyze">Проверить требования и 1С</button></div></div></div></div></div>${p.analysis?`<div class="panel"><div class="ph"><h2>Оценка реализуемости</h2>${badge(p.analysis.status)}</div><div class="body">${warnings||'<div class="warning"><b>Критических противоречий не обнаружено.</b></div>'}<h3>Предполагаемые источники 1С</h3><div class="tablewrap"><table><tr><th>Тип</th><th>Объект</th><th>Подтверждение</th><th>Основание</th></tr>${(p.analysis.candidate_objects||[]).map(x=>`<tr><td>${esc(x.type)}</td><td><b>${esc(x.name)}</b></td><td>${x.confirmed_in_snapshot?'Да':'Требует проверки'}</td><td>${esc(x.reason)}</td></tr>`).join('')}</table></div></div></div>`:''}${proto?`<div class="panel"><div class="ph"><h2>Предварительный макет</h2>${badge(proto.status)}</div><div class="body"><div class="tablewrap"><table><tr>${(proto.columns||[]).map(x=>`<th>${esc(x.name)}</th>`).join('')}</tr>${(proto.sample_rows||[]).map(r=>`<tr>${r.map(v=>`<td>${esc(v)}</td>`).join('')}</tr>`).join('')}</table></div><label>Комментарий к макету</label><textarea id="approvalComment"></textarea><div class="actions"><button class="btn accent" id="approve">Макет верный — сформировать ТЗ</button><button class="btn" id="reject">Нужно переделать</button></div></div></div>`:''}${spec?`<div class="panel"><div class="ph"><h2>Готовое техническое задание</h2>${badge(spec.status)}</div><div class="body"><p>ТЗ сформировано на основании файла, интервью, проверки реализуемости и согласованного макета.</p><div class="actions"><a class="btn primary" href="/onec-requirements/api/projects/${p.id}/spec.md" target="_blank">Открыть ТЗ</a><a class="btn" href="/onec-requirements/api/projects/${p.id}/spec" target="_blank">JSON</a></div></div></div>`:''}`;q('#saveAnswers').onclick=async()=>{const answers={};document.querySelectorAll('[data-answer]').forEach(x=>answers[x.dataset.answer]=x.value);await api(`/onec-requirements/api/projects/${id}/answers`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({answers})});openProject(id)};q('#analyze').onclick=async()=>{await q('#saveAnswers').onclick();await api(`/onec-requirements/api/projects/${id}/analyze`,{method:'POST'});openProject(id)};if(q('#approve'))q('#approve').onclick=()=>approval(true);if(q('#reject'))q('#reject').onclick=()=>approval(false);async function approval(approved){await api(`/onec-requirements/api/projects/${id}/prototype-approval`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({approved,comment:q('#approvalComment').value})});openProject(id)}}q('#newBtn').onclick=newForm;list();
</script></body></html>'''
